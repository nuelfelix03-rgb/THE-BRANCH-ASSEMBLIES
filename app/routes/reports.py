import csv
import io
from datetime import date, datetime, timedelta
from flask import Blueprint, render_template, request, Response, send_file
from flask_login import login_required
from app import db
from app.models import Member, Attendance, Ministry, Event
from app.utils.decorators import admin_required
from sqlalchemy import func, case

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
@admin_required
def index():
    today = date.today()
    first_of_month = today.replace(day=1)
    week_start = today - timedelta(days=today.weekday())

    stats = {
        'total_members': Member.query.filter_by(membership_status='Active').count(),
        'total_ministries': Ministry.query.count(),
        'today_attendance': Attendance.query.filter(
            Attendance.date == today, Attendance.status == 'Present'
        ).count(),
        'week_attendance': Attendance.query.filter(
            Attendance.date >= week_start, Attendance.date <= today,
            Attendance.status == 'Present'
        ).count(),
        'birthdays_this_month': Member.query.filter(
            db.extract('month', Member.date_of_birth) == today.month,
            Member.membership_status == 'Active'
        ).count() if hasattr(db, 'extract') else 0,
        'new_members': Member.query.filter(
            Member.date_joined >= first_of_month
        ).count(),
        'upcoming_events': Event.query.filter(
            Event.start_date >= datetime.now(), Event.status == 'Upcoming'
        ).count(),
    }

    return render_template('reports/index.html', stats=stats)


@reports_bp.route('/members')
@login_required
@admin_required
def members_report():
    total_members = Member.query.count()
    active_members = Member.query.filter_by(membership_status='Active').count()
    new_this_month = Member.query.filter(
        func.strftime('%Y-%m', Member.date_joined) == date.today().strftime('%Y-%m')
    ).count()
    by_ministry = db.session.query(
        Ministry.name, func.count(Member.id)
    ).outerjoin(Ministry, Member.ministry_id == Ministry.id
    ).group_by(Ministry.name).all()
    by_gender = db.session.query(
        Member.gender, func.count(Member.id)
    ).group_by(Member.gender).all()

    return render_template('reports/members.html',
        total=total_members, active=active_members,
        new_this_month=new_this_month,
        by_ministry=by_ministry, by_gender=by_gender)


@reports_bp.route('/attendance')
@login_required
@admin_required
def attendance_report():
    today = date.today()
    period = request.args.get('period', 'month')

    if period == 'week':
        start_date = today - timedelta(days=today.weekday())
    elif period == 'month':
        start_date = today.replace(day=1)
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today - timedelta(days=30)

    records = db.session.query(
        Attendance.service_type,
        func.count(Attendance.id).label('total'),
        func.sum(case((Attendance.status == 'Present', 1), else_=0)).label('present'),
        func.sum(case((Attendance.status == 'Absent', 1), else_=0)).label('absent')
    ).filter(Attendance.date >= start_date).group_by(Attendance.service_type).all()

    daily_totals = db.session.query(
        Attendance.date,
        func.count(Attendance.id).label('total')
    ).filter(Attendance.date >= start_date).group_by(Attendance.date).order_by(Attendance.date).all()

    return render_template('reports/attendance.html',
        records=records, daily_totals=daily_totals,
        period=period, start_date=start_date, today=today)


@reports_bp.route('/export/members/csv')
@login_required
@admin_required
def export_members_csv():
    members = Member.query.all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Name', 'Gender', 'Phone', 'Email', 'Ministry', 'Status', 'Date Joined'])
    for m in members:
        writer.writerow([m.member_id, m.full_name(), m.gender, m.phone_number,
                        m.email, m.ministry.name if m.ministry else '',
                        m.membership_status, m.date_joined])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=members_report.csv'}
    )


@reports_bp.route('/export/attendance/csv')
@login_required
@admin_required
def export_attendance_csv():
    records = Attendance.query.order_by(Attendance.date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Member', 'Service Type', 'Status', 'Recorded By'])
    for r in records:
        writer.writerow([r.date, r.member.full_name() if r.member else 'N/A',
                        r.service_type, r.status,
                        r.recorded_by_user.username if r.recorded_by_user else 'N/A'])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=attendance_report.csv'}
    )


@reports_bp.route('/export/members/pdf')
@login_required
@admin_required
def export_members_pdf():
    if not REPORTLAB_AVAILABLE:
        flash('PDF export requires reportlab. Install with: pip install reportlab', 'warning')
        return redirect(url_for('reports.members_report'))

    members = Member.query.all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Member Report', styles['Title']))
    elements.append(Spacer(1, 0.25 * inch))
    elements.append(Paragraph(f'Generated: {datetime.now().strftime("%B %d, %Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.25 * inch))

    data = [['ID', 'Name', 'Gender', 'Phone', 'Email', 'Ministry', 'Status']]
    for m in members:
        data.append([
            m.member_id, m.full_name(), m.gender or '',
            m.phone_number or '', m.email or '',
            m.ministry.name if m.ministry else '', m.membership_status
        ])

    table = Table(data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ])
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                    as_attachment=True, download_name='members_report.pdf')


@reports_bp.route('/export/members/excel')
@login_required
@admin_required
def export_members_excel():
    if not OPENPYXL_AVAILABLE:
        flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'warning')
        return redirect(url_for('reports.members_report'))

    members = Member.query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Members'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    headers = ['ID', 'Name', 'Gender', 'Phone', 'Email', 'Ministry', 'Status', 'Date Joined']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, m in enumerate(members, 2):
        ws.cell(row=row, column=1, value=m.member_id)
        ws.cell(row=row, column=2, value=m.full_name())
        ws.cell(row=row, column=3, value=m.gender or '')
        ws.cell(row=row, column=4, value=m.phone_number or '')
        ws.cell(row=row, column=5, value=m.email or '')
        ws.cell(row=row, column=6, value=m.ministry.name if m.ministry else '')
        ws.cell(row=row, column=7, value=m.membership_status)
        ws.cell(row=row, column=8, value=str(m.date_joined or ''))

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name='members_report.xlsx')


@reports_bp.route('/export/attendance/excel')
@login_required
@admin_required
def export_attendance_excel():
    if not OPENPYXL_AVAILABLE:
        flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'warning')
        return redirect(url_for('reports.attendance_report'))

    records = Attendance.query.order_by(Attendance.date.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    headers = ['Date', 'Member', 'Service Type', 'Status', 'Recorded By']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, r in enumerate(records, 2):
        ws.cell(row=row, column=1, value=str(r.date))
        ws.cell(row=row, column=2, value=r.member.full_name() if r.member else 'N/A')
        ws.cell(row=row, column=3, value=r.service_type)
        ws.cell(row=row, column=4, value=r.status)
        ws.cell(row=row, column=5, value=r.recorded_by_user.username if r.recorded_by_user else 'N/A')

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name='attendance_report.xlsx')
